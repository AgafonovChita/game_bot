import datetime

from bot.services.repo import SQLAlchemyRepo, ProtocolRepo
from openpyxl import Workbook
from openpyxl.styles import Alignment


async def export_final_protocol(repo: SQLAlchemyRepo):
    final_protol_list = await repo.get_repo(ProtocolRepo).get_final_protocol_to_export()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:D1')
    cell = worksheet["A1"]
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    worksheet.column_dimensions['A'].width = 10
    worksheet.row_dimensions[1].height = 30
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 30
    worksheet.cell(row=1, column=1, value=f"Итоговый протокол игры\n" \
                                          f"Время экспорта: {datetime.datetime.today()}")
    worksheet.cell(row=2, column=1, value="Место")
    worksheet.cell(row=2, column=2, value="Команда")
    worksheet.cell(row=2, column=3, value="Финальная точка")
    worksheet.cell(row=2, column=4, value="Финальное время")

    for row, protocol in enumerate(final_protol_list, start=1):
        worksheet.cell(row=row+2, column=1, value=row)
        worksheet.cell(row=row+2, column=2, value=protocol.team_name)
        worksheet.cell(row=row+2, column=3, value=protocol.point_id)
        worksheet.cell(row=row+2, column=4, value=protocol.finish_time)

    name_file_protocol = f"final_protocol.xlsx"
    workbook.save(name_file_protocol)
    return name_file_protocol


async def export_current_protocol(repo: SQLAlchemyRepo):
    export_protocol = await repo.get_repo(ProtocolRepo).get_current_protocol_to_export()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:D1')
    cell = worksheet["A1"]
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center" )
    worksheet.column_dimensions['A'].width = 15
    worksheet.row_dimensions[1].height = 30
    worksheet.column_dimensions['B'].width = 7
    worksheet.column_dimensions['C'].width = 40
    worksheet.cell(row=1, column=1, value=f"Экспорт протокола игры\n" \
                                          f"Дата игры: {datetime.datetime.today()}")
    worksheet.cell(row=2, column=1, value="Команда")
    worksheet.cell(row=2, column=2, value="Уровень")
    worksheet.cell(row=2, column=3, value="Время прохождения уровня")

    for row, protocol in enumerate(export_protocol, start=3):
        worksheet.cell(row=row, column=1, value=protocol.team_name)
        worksheet.cell(row=row, column=2, value=protocol.point_id)
        worksheet.cell(row=row, column=3, value=protocol.point_close_time)

    name_file_protocol = f"current_protocol.xlsx"
    workbook.save(name_file_protocol)
    return name_file_protocol
